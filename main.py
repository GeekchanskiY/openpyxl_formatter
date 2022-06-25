import re
import phonenumbers
from openpyxl import Workbook, load_workbook, worksheet
from sys import argv

# filename - input .xlsx workbook file name
# col 1-3 cols, if they do not exist input 0
args = argv
filename = ""
cols: list[int] = []

try:
    filename = args[1]
except IndexError:
    input("Filename is not provided. Press enter to exit...")
    exit()

for arg in args[2:]:
    try:
        if int(arg) != 0:
            cols.append((int(arg)-1))
    except ValueError:
        input(f"Col value {arg} seems non-numeric. Press enter to exit...")
        exit()


def check_number(string: str) -> int:
    """
        Checks if number matches one of these cases:
            Case 1: code starts with 9
            Case 2: starts with 6 & 9
        and returns if matches

    Args:
        string(str): formatted number str (length will be checked dynamically)
    Return:
        bool: 0 - does not match, 1 - case 1, 2 - case 2
    """
    if len(string) == 10:
        if string[0] == "9":
            return 1
        elif string[3] == "6":
            return 2
        elif string[3] == "9":
            return 2
    elif len(string) == 7:
        if string[0] == "6":
            return 2
        elif string[0] == "9":
            return 2
    return 0


def alternative_select_number_from_str(string: str) -> str:
    """

    Alternative string formatting for all specific cases

    Args:
         string(str): raw number str
    Return:
        str: formatted string
    """
    # deleting all unnecessary symbols

    string = re.sub("[а-я]", "", string)
    string = re.sub("[А-Я]", "", string)
    string = re.sub("[a-z]", "", string)
    string = re.sub("[A-Z]", "", string)

    # deleting spaces
    string = string.strip()

    # for (xxx) case
    if string[0] == "(" and string[-1] == ")":
        string = string[1:-1]

    # flag for getting code
    code_starts = False
    code = ""
    this_phone = ""
    for i in range(0, len(string)):

        # getting code
        if string[i] == "(":
            code_starts = True
        elif string[i] == ")":
            code_starts = False

        # if numeric value in string pos, writing this number to code or phone
        if re.match(r"\d", string[i]):
            if code_starts is True:
                code += string[i]
            else:
                this_phone += string[i]

    # concat code with phone
    this_phone = code + this_phone

    return this_phone


def write_to_new_workbook(output_data: list[dict]) -> bool:
    """
        Writes data to output.xlsx
    Args:
        output_data (list[dict]) - data
    Return:
        bool - if created
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Output"
    row = 1
    for i in output_data:
        if i["first"]:
            ws.cell(row=row, column=1).value = i["name"]
            ws.cell(row=row, column=2).value = i["phone"]
            ws.cell(row=row, column=3).value = i["comment"]
            row += 1
    ws.cell(row=row, column=1).value = "//////////////////"
    ws.cell(row=row, column=2).value = "// I <3 Python ///"
    ws.cell(row=row, column=3).value = "//////////////////"
    row += 1
    for i in output_data:
        if not i["first"]:
            ws.cell(row=row, column=1).value = i["name"]
            ws.cell(row=row, column=2).value = i["phone"]
            ws.cell(row=row, column=3).value = i["comment"]
            row += 1
    wb.save("output.xlsx")
    return True


def filetype1(wb: Workbook) -> list:
    """
        Makes new worksheet for filetype 1 and returns 1 if created successfully

        Args:
               wb (Workbook): original workbook

        Returns:
            list: list with formatted data
    """

    output_data: list[dict] = []
    ws: worksheet = wb.active
    for row in ws.values:

        raw_numbers: list[str] = []
        # row 2

        # First filter with general regex
        for col in cols:
            numbers: list[str] = []
            for m in re.finditer(r"(?:(?:8|\+7)[\- ]?)?(?:\(?\d{3}\)?[\- ]?)?[\d\- ]{7,10}", str(row[col])):

                # Second filter with phonenumbers

                for match in phonenumbers.PhoneNumberMatcher(m.string, "RU"):
                    raw_numbers.append(match.raw_string)
                    numbers.append(str(match.number.national_number))

                # deleting everything phonenumbers found

                temp_str: str = m.string

                for number in raw_numbers:
                    temp_str = temp_str.replace(number, "")

                # Third custom filter

                if len(str(re.sub(r"\D", "", temp_str))) >= 7:
                    for t_s in temp_str.split(","):
                        if t_s != "" and len(re.sub(r"\D", "", t_s)) >= 7:
                            numbers.append(alternative_select_number_from_str(t_s))

                # deleting duplicates if exists

            numbers = list(set(numbers))

            # creating output data

            for n in numbers:
                res = check_number(n)
                if res == 1:
                    comment = ""
                    for comm in cols:
                        if str(row[comm]) != "None":
                            comment = str(row[comm])
                            break
                    output_data.append({"name": str(row[0]), "phone": "7"+n, "comment": comment, "first": True})

                elif res == 2:
                    comment = ""
                    for comm in cols:
                        if str(row[comm]) != "None":
                            comment = str(row[comm])
                            break
                    if len(n) == 7:
                        n = "7812" + n
                        output_data.append({"name": str(row[0]), "phone": n, "comment": comment,
                                            "first": False})
                    else:
                        output_data.append({"name": str(row[0]), "phone": "7"+n, "comment": comment,
                                            "first": False})
    return output_data


def filetype2(wb: Workbook) -> list:
    """
        Makes new worksheet for filetype 1 and returns 1 if created successfully

        Args:
               wb (Workbook): original workbook

        Returns:
            list: list with formatted data
    """
    output_data: list[dict] = []
    ws: worksheet = wb.active
    last_found = ""
    for row in ws.values:

        raw_numbers: list[str] = []
        if str(row[0]) != "None" and str(row[0]) != "" and str(row[0]) != " ":
            last_found = str(row[0])
            print(row[0])

        # First filter with general regex

        for col in cols:
            numbers: list[str] = []
            for m in re.finditer(r"(?:(?:8|\+7)[\- ]?)?(?:\(?\d{3}\)?[\- ]?)?[\d\- ]{7,10}", str(row[col])):

                # Second filter with phonenumbers

                for match in phonenumbers.PhoneNumberMatcher(m.string, "RU"):
                    raw_numbers.append(match.raw_string)
                    numbers.append(str(match.number.national_number))

                # deleting everything phonenumbers found

                temp_str: str = m.string

                for number in raw_numbers:
                    temp_str = temp_str.replace(number, "")

                # Third custom filter

                if len(str(re.sub(r"\D", "", temp_str))) >= 7:
                    for t_s in temp_str.split(","):
                        if t_s != "" and len(re.sub(r"\D", "", t_s)) >= 7:
                            numbers.append(alternative_select_number_from_str(t_s))

                # deleting duplicates if exists

            numbers = list(set(numbers))

            # creating output data
            for n in numbers:
                res = check_number(n)
                if res == 1:
                    comment = ""
                    for comm in cols:
                        if str(row[comm]) != "None":
                            comment = str(row[comm])
                            break
                    output_data.append({"name": last_found, "phone": "7"+n, "comment": comment, "first": True})

                elif res == 2:
                    comment = ""
                    for comm in cols:
                        if str(row[comm]) != "None":
                            comment = str(row[comm])
                            break
                    if len(n) == 7:
                        n = "7812" + n
                        output_data.append({"name": last_found, "phone": n, "comment": comment,
                                            "first": False})
                    else:
                        output_data.append({"name": last_found, "phone": "7" + n, "comment": comment,
                                            "first": False})

    return output_data


def main(file: str):
    wb = load_workbook(file)
    ws = wb.active
    row_len = len(list(ws.iter_rows())[0])
    if row_len == 3:
        data = filetype1(wb)
    else:
        data = filetype2(wb)
    write_to_new_workbook(data)


if __name__ == '__main__':
    main(filename)
